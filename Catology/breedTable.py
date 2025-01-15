import openpyxl
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook

def count_values_per_attribute():
    file_path = 'static/final4.xlsx'
    data = pd.read_excel(file_path)


    breeds_id = data.iloc[:, 1].unique()

    print(breeds_id)
    not_good_i = [0, 1, 3, 8, 10, 14, 15, 17, 18, 19, 20, 22]

    all_att = []
    for breed in range(1, 13):
        filtered_data = data[data.iloc[:, 1] == breed]  # Selectăm rândurile unde coloana 2 este 1
        att = []
        for i in range(0,25):
            if i not in not_good_i:
                count_of_zeros = (filtered_data.iloc[:, i] == 0).sum()
                count_of_ones = (filtered_data.iloc[:, i] == 1).sum()
                count_of_twos = (filtered_data.iloc[:, i] == 2).sum()
                count_of_threes = (filtered_data.iloc[:, i] == 3).sum()
                count_of_fours = (filtered_data.iloc[:, i] == 4).sum()
                # Afișăm rezultatele

                denominator = (count_of_zeros + count_of_ones + count_of_twos + count_of_threes + count_of_fours)

                if denominator == 0:
                    avg = 0  # Handle the case when no values exist
                else:
                    avg = ((
                                       count_of_zeros + count_of_ones + 2 * count_of_twos + 3 * count_of_threes + 4 * count_of_fours) /
                           denominator)

                # print(avg)
                att.append(int(avg))
        all_att.append(att)
    print(all_att)

    headers = ["Number", "Ext", "Shy", "Calm", "Scared", "Vigilant", "Affectionate", "Friendly", "Solitary", "Aggressive", "PredatorMammal", "Coat", "Intelligence_Score"]
    print("aa")
    wb = Workbook()
    ws = wb.active

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)


    transposed_matrix = list(zip(*all_att))

    for col_idx, column in enumerate(transposed_matrix, start=1):
        for row_idx, value in enumerate(column, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Salvarea fișierului Excel
    wb.save("output.xlsx")

    print("Matricea a fost salvată în fișierul 'output.xlsx'.")
# Apelarea funcției

count_values_per_attribute()